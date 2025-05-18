from openpyxl import load_workbook
import xlrd   
import csv
from PIL import Image, UnidentifiedImageError
import pytesseract
import pdfplumber
from docx import Document
import io

def extract_text_from_pdf(file_stream: io.BytesIO) -> str:
    try:
        with pdfplumber.open(file_stream) as pdf:
            return "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception:
        raise ValueError("Error reading PDF file")


def extract_text_from_docx(file_stream: io.BytesIO) -> str:
    try:
        doc = Document(file_stream)
        return "\n".join(para.text for para in doc.paragraphs)
    except Exception:
        raise ValueError("Error reading DOCX file")


def extract_text_from_txt(file_stream: io.BytesIO) -> str:
    try:
        return file_stream.read().decode("utf-8")
    except Exception:
        raise ValueError("Error reading TXT file")


def extract_text_from_image(file_storage):
    try:
        image = Image.open(file_storage.stream)
        return pytesseract.image_to_string(image).strip()
    except UnidentifiedImageError:
        raise ValueError("Invalid image file")

def extract_text_from_csv(file_stream):
    try:
        file_stream.seek(0)  # ensure we're at the beginning of the stream
        reader = csv.reader(io.StringIO(file_stream.read().decode("utf-8")))
        text = []
        for row in reader:
            line = ' '.join(cell for cell in row if cell.strip())
            if line:
                text.append(line)
        return "\n".join(text)
    except Exception:
        raise ValueError("Error reading CSV file")

def extract_text_from_xlsx(file_stream):
    try:
        file_stream.seek(0)
        wb = load_workbook(filename=file_stream, data_only=True)
        text = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                line = ' '.join(str(cell) for cell in row if cell is not None)
                if line:
                    text.append(line)
        return "\n".join(text)
    except Exception:
        raise ValueError("Error reading XLSX file")

def extract_text_from_xls(file_stream):
    try:
        file_stream.seek(0)
        book = xlrd.open_workbook(file_contents=file_stream.read())
        text = []
        for sheet in book.sheets():
            for row_idx in range(sheet.nrows):
                row = sheet.row(row_idx)
                line = ' '.join(str(cell.value) for cell in row if cell.value)
                if line:
                    text.append(line)
        return "\n".join(text)
    except Exception:
        raise ValueError("Error reading XLS file")